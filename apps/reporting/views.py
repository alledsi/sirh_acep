"""Vues du module Reporting — Sprint 5."""
import json
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import View

from apps.attendance.models import Anomaly
from apps.core.mixins import GlobalAccessRequiredMixin
from apps.employees.models import Employee

from .forms import AnomalyValidateForm
from .services import (
    compute_hours_per_employee, compute_period_stats, compute_today_status,
    get_anomalies_for_user, get_anomaly_breakdown, get_breakdown_by_agence,
    get_breakdown_by_direction, get_directeur_directions, get_directeur_employees,
    get_global_overview, get_presence_30_days,
)


class DirecteurRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Accès réservé aux Directeurs (ou RH/DG, qui peuvent aussi voir)."""

    def test_func(self):
        u = self.request.user
        return u.is_authenticated and (u.is_directeur or u.has_global_access)


# ============ Directeur — Tableau de bord ============

class DirecteurDashboardView(DirecteurRequiredMixin, View):
    template_name = 'reporting/directeur_dashboard.html'

    def get(self, request):
        directions = get_directeur_directions(request.user)
        if not directions.exists():
            messages.warning(
                request,
                "Aucune direction ne vous est rattachée. Vérifiez votre fiche employé "
                "(champ Direction) ou demandez à la RH de vous désigner directeur d'une direction."
            )
            return redirect('core:home')

        employees = get_directeur_employees(request.user)
        today_status = compute_today_status(employees)

        # Anomalies en attente
        anomalies_pending = get_anomalies_for_user(request.user, only_pending=True)[:10]
        anomalies_pending_count = get_anomalies_for_user(request.user, only_pending=True).count()

        # Stats des 30 derniers jours
        today = timezone.localdate()
        d30 = today - timedelta(days=30)
        period_stats = compute_period_stats(employees, d30, today)

        # Histogramme : heures par employé sur le mois en cours
        import json
        month_start = today.replace(day=1)
        hours_per_employee = compute_hours_per_employee(employees, month_start, today)

        return render(request, self.template_name, {
            'directions': directions,
            'employees': employees,
            'today_status': today_status,
            'today_entries_by_employee': {e.employee_id: e for e in today_status['today_entries']},
            'anomalies_pending': anomalies_pending,
            'anomalies_pending_count': anomalies_pending_count,
            'period_stats': period_stats,
            'hours_per_employee_data': json.dumps(hours_per_employee),
            'hours_per_employee_count': len(hours_per_employee),
        })


class DirecteurEquipeView(DirecteurRequiredMixin, View):
    template_name = 'reporting/directeur_equipe.html'

    def get(self, request):
        employees = get_directeur_employees(request.user)
        today = timezone.localdate()
        month_start = today.replace(day=1)

        # Stats par employé (mois en cours)
        from apps.attendance.models import TimeEntry
        employee_stats = []
        for emp in employees:
            entries = TimeEntry.objects.filter(
                employee=emp, work_date__gte=month_start, work_date__lte=today,
            )
            hours = timedelta()
            days = 0
            retards = 0
            anomalies = 0
            for e in entries:
                d = e.worked_duration
                if d:
                    hours += d
                    days += 1
                for a in e.anomalies.all():
                    anomalies += 1
                    if a.anomaly_type == Anomaly.TYPE_LATE:
                        retards += 1
            total_min = int(hours.total_seconds() // 60)
            h, m = divmod(total_min, 60)
            employee_stats.append({
                'employee': emp,
                'hours': f"{h}h{m:02d}",
                'days': days,
                'retards': retards,
                'anomalies': anomalies,
            })

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(employee_stats, 30)
        page_obj = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'employee_stats': page_obj.object_list,
            'employee_stats_count': paginator.count,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'directions': get_directeur_directions(request.user),
        })


class DirecteurAnomaliesView(DirecteurRequiredMixin, View):
    template_name = 'reporting/anomaly_list.html'

    def get(self, request):
        from django.core.paginator import Paginator
        only_pending = request.GET.get('statut', 'pending') == 'pending'
        anomalies_qs = get_anomalies_for_user(request.user, only_pending=only_pending)
        paginator = Paginator(anomalies_qs, 30)
        page_obj = paginator.get_page(request.GET.get('page'))
        return render(request, self.template_name, {
            'anomalies': page_obj.object_list,
            'anomalies_count': paginator.count,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'only_pending': only_pending,
            'scope_label': 'ma direction' if request.user.is_directeur and not request.user.has_global_access else 'toutes les directions',
        })


class AnomalyValidateView(DirecteurRequiredMixin, View):
    template_name = 'reporting/anomaly_validate.html'

    def _get_anomaly(self, request, pk):
        anomaly = get_object_or_404(Anomaly, pk=pk)
        # Vérifier que l'utilisateur a le droit
        allowed_ids = list(get_anomalies_for_user(request.user, only_pending=False).values_list('pk', flat=True))
        if anomaly.pk not in allowed_ids:
            raise Http404("Anomalie non visible pour vous.")
        return anomaly

    def get(self, request, pk):
        anomaly = self._get_anomaly(request, pk)
        form = AnomalyValidateForm()
        return render(request, self.template_name, {'anomaly': anomaly, 'form': form})

    def post(self, request, pk):
        anomaly = self._get_anomaly(request, pk)
        form = AnomalyValidateForm(request.POST)
        if form.is_valid():
            anomaly.is_acknowledged = True
            anomaly.acknowledged_by = request.user
            anomaly.acknowledged_at = timezone.now()
            anomaly.acknowledgement_note = form.cleaned_data['note']
            anomaly.save()
            messages.success(request, "Anomalie validée.")
            # Redirige vers la liste d'anomalies appropriée
            if request.user.is_directeur and not request.user.has_global_access:
                return redirect('reporting:directeur_anomalies')
            return redirect('reporting:anomaly_list')
        return render(request, self.template_name, {'anomaly': anomaly, 'form': form})


# ============ RH/DG — Statistiques globales ============

class RHStatsView(GlobalAccessRequiredMixin, View):
    template_name = 'reporting/rh_stats.html'

    def get(self, request):
        overview = get_global_overview()
        return render(request, self.template_name, {
            'overview': overview,
            'breakdown_agence': json.dumps(get_breakdown_by_agence()),
            'breakdown_direction': json.dumps(get_breakdown_by_direction()),
            'anomaly_breakdown': json.dumps(get_anomaly_breakdown()),
            'presence_30': json.dumps(get_presence_30_days()),
        })


class AnomalyListView(GlobalAccessRequiredMixin, View):
    """Liste des anomalies pour la RH/DG (toutes les anomalies)."""
    template_name = 'reporting/anomaly_list.html'

    def get(self, request):
        from django.core.paginator import Paginator
        only_pending = request.GET.get('statut', 'pending') == 'pending'
        anomalies_qs = get_anomalies_for_user(request.user, only_pending=only_pending)
        paginator = Paginator(anomalies_qs, 30)
        page_obj = paginator.get_page(request.GET.get('page'))
        return render(request, self.template_name, {
            'anomalies': page_obj.object_list,
            'anomalies_count': paginator.count,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'only_pending': only_pending,
            'scope_label': 'toutes les directions',
        })


# ============ Suivi quotidien (RH/DG) ============

class DailyTrackingView(GlobalAccessRequiredMixin, View):
    """Suivi des pointages : jour / mois / année / période libre.

    - jour  : liste détaillée (arrivée, départ, IPs, statut)
    - mois / annee / periode : agrégé par employé (jours présents, absents,
      total heures, IPs distinctes utilisées)
    """
    template_name = 'reporting/daily_tracking.html'

    PERIOD_CHOICES = [
        ('jour', 'Jour'),
        ('mois', 'Mois'),
        ('annee', 'Année'),
        ('periode', 'Période libre'),
    ]

    def _resolve_period(self, request):
        """Retourne (period_type, start, end, label)."""
        from datetime import datetime, date as _date
        period = request.GET.get('period', 'jour')
        today = timezone.localdate()

        if period == 'mois':
            year = int(request.GET.get('year') or today.year)
            month = int(request.GET.get('month') or today.month)
            start = _date(year, month, 1)
            end = (_date(year + 1, 1, 1) - timedelta(days=1)
                   if month == 12 else _date(year, month + 1, 1) - timedelta(days=1))
            label = f'{start:%B %Y}'
        elif period == 'annee':
            year = int(request.GET.get('year') or today.year)
            start, end = _date(year, 1, 1), _date(year, 12, 31)
            label = f'Année {year}'
        elif period == 'periode':
            try:
                start = datetime.strptime(request.GET.get('start_date', ''), '%Y-%m-%d').date()
            except ValueError:
                start = today - timedelta(days=7)
            try:
                end = datetime.strptime(request.GET.get('end_date', ''), '%Y-%m-%d').date()
            except ValueError:
                end = today
            if end < start:
                start, end = end, start
            label = f'Du {start:%d/%m/%Y} au {end:%d/%m/%Y}'
        else:  # jour
            try:
                start = datetime.strptime(request.GET.get('date', ''), '%Y-%m-%d').date()
            except ValueError:
                start = today
            end = start
            period = 'jour'
            label = f'{start:%A %d %B %Y}'
        return period, start, end, label

    def get(self, request):
        from datetime import date as _date, datetime
        from apps.attendance.models import TimeEntry
        from apps.employees.models import Employee
        from apps.organization.models import Agence, Direction, Mutuelle

        period, start, end, label = self._resolve_period(request)

        mutuelle_id = request.GET.get('mutuelle', '')
        agence_id = request.GET.get('agence', '')
        direction_id = request.GET.get('direction', '')

        employees = Employee.objects.filter(is_active=True).select_related(
            'user', 'bureau__agence__mutuelle', 'direction',
        )
        if mutuelle_id:
            employees = employees.filter(bureau__agence__mutuelle_id=mutuelle_id)
        if agence_id:
            employees = employees.filter(bureau__agence_id=agence_id)
        if direction_id:
            employees = employees.filter(direction_id=direction_id)

        entries_qs = TimeEntry.objects.filter(
            work_date__gte=start, work_date__lte=end,
            employee__in=employees,
        ).select_related('employee__user', 'arrival_bureau', 'departure_bureau')

        rows = []
        present_count = 0
        absent_count = 0
        total_minutes_all = 0

        if period == 'jour':
            # Mode jour : ligne par employé, statut + IPs de la journée
            entries_by_emp = {e.employee_id: e for e in entries_qs}
            for emp in employees:
                entry = entries_by_emp.get(emp.pk)
                if entry and entry.arrival_time:
                    status = 'departed' if entry.departure_time else ('on_break' if entry.is_on_break else 'present')
                    present_count += 1
                else:
                    status = 'absent'
                    absent_count += 1
                rows.append({'employee': emp, 'entry': entry, 'status': status})
            order = {'present': 0, 'on_break': 1, 'departed': 2, 'absent': 3}
            rows.sort(key=lambda r: (order[r['status']], r['employee'].user.matricule))
        else:
            # Mode agrégé : compte les jours présents/absents sur la période
            entries_by_emp_map = {}
            for e in entries_qs:
                entries_by_emp_map.setdefault(e.employee_id, []).append(e)

            # Calcul du nombre de jours ouvrés dans la période (Lun-Sam par défaut)
            total_workable_days = sum(
                1 for i in range((end - start).days + 1)
                if (start + timedelta(days=i)).weekday() < 6  # exclut dimanche
            )

            for emp in employees:
                emp_entries = entries_by_emp_map.get(emp.pk, [])
                days_present = sum(1 for e in emp_entries if e.arrival_time)
                total = timedelta()
                arrival_ips = set()
                departure_ips = set()
                for e in emp_entries:
                    if e.arrival_ip:
                        arrival_ips.add(e.arrival_ip)
                    if e.departure_ip:
                        departure_ips.add(e.departure_ip)
                    d = e.worked_duration
                    if d:
                        total += d
                minutes = int(total.total_seconds() // 60)
                total_minutes_all += minutes
                h, m = divmod(minutes, 60)
                days_absent = max(0, total_workable_days - days_present)
                if days_present:
                    present_count += 1
                else:
                    absent_count += 1
                rows.append({
                    'employee': emp,
                    'days_present': days_present,
                    'days_absent': days_absent,
                    'hours': f'{h}h{m:02d}',
                    'minutes': minutes,
                    'arrival_ips': sorted(arrival_ips),
                    'departure_ips': sorted(departure_ips),
                })
            rows.sort(key=lambda r: r['minutes'], reverse=True)

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(rows, 50)
        page_obj = paginator.get_page(request.GET.get('page'))

        h_all, m_all = divmod(total_minutes_all, 60)

        return render(request, self.template_name, {
            'period': period,
            'start': start,
            'end': end,
            'period_label': label,
            'target_date': start,  # rétro-compat avec le template
            'rows': page_obj.object_list,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'present_count': present_count,
            'absent_count': absent_count,
            'total_count': len(rows),
            'total_hours_all': f'{h_all}h{m_all:02d}',
            'mutuelles': Mutuelle.objects.filter(is_active=True),
            'agences': Agence.objects.filter(is_active=True).select_related('mutuelle'),
            'directions': Direction.objects.filter(is_active=True),
            'filters': {
                'mutuelle': mutuelle_id, 'agence': agence_id, 'direction': direction_id,
            },
            'year': start.year,
            'month': start.month,
            'years': list(range(timezone.localdate().year - 3, timezone.localdate().year + 1)),
            'months': [
                (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
                (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
                (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre'),
            ],
            'period_choices': self.PERIOD_CHOICES,
        })


# ============ Cumul horaire mensuel (RH/DG) ============

class ExportStatsView(GlobalAccessRequiredMixin, View):
    """Export Excel des statistiques (jour / semaine / mois / année).

    URL : ?type=daily&date=YYYY-MM-DD
          ?type=weekly&date=YYYY-MM-DD       (semaine contenant cette date)
          ?type=monthly&year=YYYY&month=MM
          ?type=yearly&year=YYYY
    """

    def _export_daily_full(self, request, start, end, label, period):
        """Export Excel : présents + absents + IPs sur la période choisie.

        En mode 'jour' : 1 ligne par employé (pointage détaillé).
        En mode agrégé (mois/année/période) : cumul par employé.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from django.http import HttpResponse
        from apps.attendance.models import TimeEntry
        from apps.employees.models import Employee

        # Filtres appliqués (repris de l'écran Suivi)
        mutuelle_id = request.GET.get('mutuelle', '')
        agence_id = request.GET.get('agence', '')
        direction_id = request.GET.get('direction', '')
        employees = Employee.objects.filter(is_active=True).select_related(
            'user', 'bureau__agence__mutuelle', 'direction',
        )
        if mutuelle_id:
            employees = employees.filter(bureau__agence__mutuelle_id=mutuelle_id)
        if agence_id:
            employees = employees.filter(bureau__agence_id=agence_id)
        if direction_id:
            employees = employees.filter(direction_id=direction_id)

        entries_qs = TimeEntry.objects.filter(
            work_date__gte=start, work_date__lte=end, employee__in=employees,
        ).select_related('arrival_bureau', 'departure_bureau')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Suivi'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='02564A', end_color='02564A', fill_type='solid')
        title_font = Font(bold=True, size=14, color='02564A')
        red_fill = PatternFill(start_color='FDECEC', end_color='FDECEC', fill_type='solid')

        ws['A1'] = f'ACEP — {label} ({employees.count()} agent(s) filtré(s))'
        ws['A1'].font = title_font

        if period == 'jour':
            headers = ['Matricule', 'Nom complet', 'Fonction', 'Direction', 'Bureau',
                       'Arrivée', 'IP arrivée',
                       'Départ', 'IP départ',
                       'Heures travaillées', 'Statut']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center')

            entries_by_emp = {e.employee_id: e for e in entries_qs}
            row_num = 4
            for emp in employees:
                entry = entries_by_emp.get(emp.pk)
                is_absent = not (entry and entry.arrival_time)
                if entry and entry.worked_duration:
                    total_min = int(entry.worked_duration.total_seconds() // 60)
                    h, m = divmod(total_min, 60)
                    hours_str = f'{h}h{m:02d}'
                else:
                    hours_str = ''
                if is_absent:
                    status = 'ABSENT'
                elif entry.departure_time:
                    status = 'Parti'
                elif entry.is_on_break:
                    status = 'En pause'
                else:
                    status = 'Présent'

                values = [
                    emp.user.matricule,
                    emp.user.get_full_name() or emp.user.matricule,
                    emp.position or '',
                    emp.direction.name if emp.direction else '',
                    emp.bureau.name if emp.bureau else '',
                    entry.arrival_time.strftime('%H:%M') if entry and entry.arrival_time else '',
                    entry.arrival_ip if entry and entry.arrival_ip else '',
                    entry.departure_time.strftime('%H:%M') if entry and entry.departure_time else '',
                    entry.departure_ip if entry and entry.departure_ip else '',
                    hours_str,
                    status,
                ]
                for col, v in enumerate(values, 1):
                    c = ws.cell(row=row_num, column=col, value=v)
                    if is_absent:
                        c.fill = red_fill
                row_num += 1

            widths = [12, 30, 24, 32, 25, 10, 16, 10, 16, 16, 12]
        else:
            headers = ['Matricule', 'Nom complet', 'Fonction', 'Direction', 'Bureau',
                       'Jours pointés', 'Jours absents',
                       'Total heures', 'IPs arrivée (distinctes)', 'IPs départ (distinctes)']
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center')

            entries_map = {}
            for e in entries_qs:
                entries_map.setdefault(e.employee_id, []).append(e)

            # Nb jours travaillables sur la période (exclut dimanches)
            total_workable = sum(
                1 for i in range((end - start).days + 1)
                if (start + timedelta(days=i)).weekday() < 6
            )

            row_num = 4
            for emp in employees:
                emp_entries = entries_map.get(emp.pk, [])
                days_present = sum(1 for e in emp_entries if e.arrival_time)
                days_absent = max(0, total_workable - days_present)
                total = timedelta()
                arrival_ips = set()
                departure_ips = set()
                for e in emp_entries:
                    if e.arrival_ip:
                        arrival_ips.add(e.arrival_ip)
                    if e.departure_ip:
                        departure_ips.add(e.departure_ip)
                    d = e.worked_duration
                    if d:
                        total += d
                total_min = int(total.total_seconds() // 60)
                h, m = divmod(total_min, 60)
                is_absent_all = days_present == 0

                values = [
                    emp.user.matricule,
                    emp.user.get_full_name() or emp.user.matricule,
                    emp.position or '',
                    emp.direction.name if emp.direction else '',
                    emp.bureau.name if emp.bureau else '',
                    days_present,
                    days_absent,
                    f'{h}h{m:02d}',
                    ', '.join(sorted(arrival_ips)),
                    ', '.join(sorted(departure_ips)),
                ]
                for col, v in enumerate(values, 1):
                    c = ws.cell(row=row_num, column=col, value=v)
                    if is_absent_all:
                        c.fill = red_fill
                row_num += 1
            widths = [12, 30, 24, 32, 25, 14, 14, 14, 30, 30]

        for col, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'suivi_acep_{period}_{start:%Y%m%d}_{end:%Y%m%d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def get(self, request):
        from datetime import datetime, date as _date
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from django.http import HttpResponse
        from apps.attendance.models import TimeEntry
        from apps.employees.models import Employee

        export_type = request.GET.get('type', 'monthly')
        today = timezone.localdate()

        # Détermination de la période
        if export_type == 'daily':
            d_str = request.GET.get('date', today.isoformat())
            d = datetime.strptime(d_str, '%Y-%m-%d').date()
            start, end, label = d, d, f'Pointages du {d:%d/%m/%Y}'
        elif export_type == 'daily_full':
            # Suivi complet — présents + absents + IPs. Prend les mêmes params
            # que la page Suivi quotidien (period=jour/mois/annee/periode).
            period = request.GET.get('period', 'jour')
            if period == 'mois':
                year = int(request.GET.get('year') or today.year)
                month = int(request.GET.get('month') or today.month)
                start = _date(year, month, 1)
                end = (_date(year + 1, 1, 1) - timedelta(days=1)
                       if month == 12 else _date(year, month + 1, 1) - timedelta(days=1))
                label = f'Suivi — {start:%B %Y}'
            elif period == 'annee':
                year = int(request.GET.get('year') or today.year)
                start, end = _date(year, 1, 1), _date(year, 12, 31)
                label = f'Suivi — Année {year}'
            elif period == 'periode':
                try:
                    start = datetime.strptime(request.GET.get('start_date', ''), '%Y-%m-%d').date()
                except ValueError:
                    start = today - timedelta(days=7)
                try:
                    end = datetime.strptime(request.GET.get('end_date', ''), '%Y-%m-%d').date()
                except ValueError:
                    end = today
                if end < start:
                    start, end = end, start
                label = f'Suivi du {start:%d/%m/%Y} au {end:%d/%m/%Y}'
            else:
                try:
                    start = datetime.strptime(request.GET.get('date', ''), '%Y-%m-%d').date()
                except ValueError:
                    start = today
                end = start
                label = f'Suivi du {start:%d/%m/%Y}'
            return self._export_daily_full(request, start, end, label, period)
        elif export_type == 'weekly':
            d_str = request.GET.get('date', today.isoformat())
            d = datetime.strptime(d_str, '%Y-%m-%d').date()
            monday = d - timedelta(days=d.weekday())
            sunday = monday + timedelta(days=6)
            start, end, label = monday, sunday, f'Semaine du {monday:%d/%m} au {sunday:%d/%m/%Y}'
        elif export_type == 'yearly':
            year = int(request.GET.get('year') or today.year)
            start, end, label = _date(year, 1, 1), _date(year, 12, 31), f'Année {year}'
        else:  # monthly
            year = int(request.GET.get('year') or today.year)
            month = int(request.GET.get('month') or today.month)
            start = _date(year, month, 1)
            if month == 12:
                end = _date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end = _date(year, month + 1, 1) - timedelta(days=1)
            label = f'{start:%B %Y}'

        # Génération du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pointages'

        # En-tête
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='02564A', end_color='02564A', fill_type='solid')
        title_font = Font(bold=True, size=14, color='02564A')

        ws['A1'] = f'ACEP — Pointages — {label}'
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Matricule', 'Nom complet', 'Direction', 'Bureau', 'Date',
                   'Arrivée', 'Départ', 'Heures travaillées']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Données
        entries = (
            TimeEntry.objects.filter(work_date__gte=start, work_date__lte=end)
            .select_related('employee__user', 'employee__bureau', 'employee__direction')
            .order_by('work_date', 'employee__user__matricule')
        )

        row_num = 4
        for e in entries:
            d = e.worked_duration
            hours_str = ''
            if d:
                total_min = int(d.total_seconds() // 60)
                h, m = divmod(total_min, 60)
                hours_str = f'{h}h{m:02d}'
            ws.cell(row=row_num, column=1, value=e.employee.user.matricule)
            ws.cell(row=row_num, column=2, value=e.employee.user.get_full_name() or e.employee.user.matricule)
            ws.cell(row=row_num, column=3, value=e.employee.direction.name if e.employee.direction else '')
            ws.cell(row=row_num, column=4, value=e.employee.bureau.name if e.employee.bureau else '')
            ws.cell(row=row_num, column=5, value=e.work_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=6, value=e.arrival_time.strftime('%H:%M') if e.arrival_time else '')
            ws.cell(row=row_num, column=7, value=e.departure_time.strftime('%H:%M') if e.departure_time else '')
            ws.cell(row=row_num, column=8, value=hours_str)
            row_num += 1

        # Largeur des colonnes
        widths = [12, 30, 25, 20, 12, 10, 10, 16]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        # 2ème feuille : Cumul par employé
        ws2 = wb.create_sheet('Cumul par employé')
        ws2['A1'] = f'Cumul des heures — {label}'
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:F1')
        ws2['A1'].alignment = Alignment(horizontal='center')

        headers2 = ['Matricule', 'Nom complet', 'Fonction', 'Direction', 'Jours travaillés', 'Total heures']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        employees = Employee.objects.filter(is_active=True).select_related('user', 'direction')
        row_num = 4
        for emp in employees:
            entries_e = TimeEntry.objects.filter(
                employee=emp, work_date__gte=start, work_date__lte=end,
            )
            total = timedelta()
            days = 0
            for e in entries_e:
                wd = e.worked_duration
                if wd:
                    total += wd
                    days += 1
            if days == 0:
                continue
            total_min = int(total.total_seconds() // 60)
            h, m = divmod(total_min, 60)
            ws2.cell(row=row_num, column=1, value=emp.user.matricule)
            ws2.cell(row=row_num, column=2, value=emp.user.get_full_name() or emp.user.matricule)
            ws2.cell(row=row_num, column=3, value=emp.position or '')
            ws2.cell(row=row_num, column=4, value=emp.direction.name if emp.direction else '')
            ws2.cell(row=row_num, column=5, value=days)
            ws2.cell(row=row_num, column=6, value=f'{h}h{m:02d}')
            row_num += 1

        for col, w in enumerate([12, 30, 24, 25, 16, 14], 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'pointages_acep_{export_type}_{start:%Y%m%d}_{end:%Y%m%d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class MonthlyHoursView(GlobalAccessRequiredMixin, View):
    """Récap mensuel des heures pointées par agent."""
    template_name = 'reporting/monthly_hours.html'

    def get(self, request):
        from datetime import date as _date
        from apps.attendance.models import TimeEntry
        from apps.employees.models import Employee
        from .services import _fmt_hm

        # Mois sélectionné (par défaut : mois en cours)
        year = int(request.GET.get('year') or timezone.localdate().year)
        month = int(request.GET.get('month') or timezone.localdate().month)
        # Bornes
        start = _date(year, month, 1)
        if month == 12:
            end = _date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end = _date(year, month + 1, 1) - timedelta(days=1)

        # Filtres
        mutuelle_id = request.GET.get('mutuelle', '')
        direction_id = request.GET.get('direction', '')

        employees = Employee.objects.filter(is_active=True).select_related(
            'user', 'bureau__agence__mutuelle', 'direction',
        )
        if mutuelle_id:
            employees = employees.filter(bureau__agence__mutuelle_id=mutuelle_id)
        if direction_id:
            employees = employees.filter(direction_id=direction_id)

        rows = []
        total_minutes_all = 0
        for emp in employees:
            entries = TimeEntry.objects.filter(
                employee=emp, work_date__gte=start, work_date__lte=end,
            )
            total = timedelta()
            days = 0
            for e in entries:
                d = e.worked_duration
                if d:
                    total += d
                    days += 1
            minutes = int(total.total_seconds() // 60)
            total_minutes_all += minutes
            rows.append({
                'employee': emp,
                'hours': _fmt_hm(total),
                'days': days,
                'minutes': minutes,
            })

        rows.sort(key=lambda r: r['minutes'], reverse=True)
        rows_count = len(rows)

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(rows, 50)
        page_obj = paginator.get_page(request.GET.get('page'))

        h_all, m_all = divmod(total_minutes_all, 60)
        from apps.organization.models import Direction, Mutuelle
        return render(request, self.template_name, {
            'year': year,
            'month': month,
            'start': start,
            'end': end,
            'rows': page_obj.object_list,
            'rows_count': rows_count,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'total_hours_all': f'{h_all}h{m_all:02d}',
            'mutuelles': Mutuelle.objects.filter(is_active=True),
            'directions': Direction.objects.filter(is_active=True),
            'filters': {'mutuelle': mutuelle_id, 'direction': direction_id},
            'years': list(range(timezone.localdate().year - 3, timezone.localdate().year + 1)),
            'months': [
                (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
                (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre'),
            ],
        })
