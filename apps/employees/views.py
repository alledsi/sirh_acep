"""Vues CRUD du module Employees (réservées RH/DG) + vue Mon profil."""
from datetime import date, datetime, timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q as models_q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView, DeleteView, DetailView, FormView, ListView, UpdateView, View,
)

from apps.core.mixins import GlobalAccessRequiredMixin
from apps.core.models import User
from apps.organization.models import Agence, Bureau, Direction, Mutuelle

from .forms import (
    EmployeeCreateForm, EmployeeDocumentForm, EmployeeUpdateForm,
)
from .models import Employee, EmployeeDocument


# ============ Liste + détail ============

class EmployeeListView(GlobalAccessRequiredMixin, ListView):
    model = Employee
    template_name = 'employees/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 30

    def get_queryset(self):
        qs = (
            Employee.objects
            .select_related('user', 'bureau__agence__mutuelle', 'direction', 'manager__user')
        )
        q = self.request.GET.get('q', '').strip()
        mutuelle = self.request.GET.get('mutuelle')
        agence = self.request.GET.get('agence')
        bureau = self.request.GET.get('bureau')
        direction = self.request.GET.get('direction')
        role = self.request.GET.get('role')
        statut = self.request.GET.get('statut', 'all')

        if q:
            qs = qs.filter(
                models_q(user__matricule__icontains=q) |
                models_q(user__first_name__icontains=q) |
                models_q(user__last_name__icontains=q) |
                models_q(user__email__icontains=q) |
                models_q(position__icontains=q)
            )
        if mutuelle:
            qs = qs.filter(bureau__agence__mutuelle_id=mutuelle)
        if agence:
            qs = qs.filter(bureau__agence_id=agence)
        if bureau:
            qs = qs.filter(bureau_id=bureau)
        if direction:
            qs = qs.filter(direction_id=direction)
        if role:
            qs = qs.filter(user__roles__contains=[role])
        if statut == 'active':
            qs = qs.filter(is_active=True)
        elif statut == 'inactive':
            qs = qs.filter(is_active=False)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['mutuelles'] = Mutuelle.objects.filter(is_active=True)
        ctx['agences'] = Agence.objects.filter(is_active=True).select_related('mutuelle')
        ctx['bureaux'] = Bureau.objects.filter(is_active=True).select_related('agence')
        ctx['directions'] = Direction.objects.filter(is_active=True)
        ctx['filters'] = {
            'q': self.request.GET.get('q', ''),
            'mutuelle': self.request.GET.get('mutuelle', ''),
            'agence': self.request.GET.get('agence', ''),
            'bureau': self.request.GET.get('bureau', ''),
            'direction': self.request.GET.get('direction', ''),
            'role': self.request.GET.get('role', ''),
            'statut': self.request.GET.get('statut', 'all'),
        }
        from apps.core.models import User
        ctx['role_choices'] = User.ROLE_CHOICES
        return ctx


class EmployeeDetailView(LoginRequiredMixin, DetailView):
    """Détail d'un employé.

    Accès : RH/DG (vue globale) OU un directeur pour les employés de sa
    direction (et sous-directions récursives). Les autres → 403.
    """
    model = Employee
    template_name = 'employees/employee_detail.html'
    context_object_name = 'employee'

    def get_queryset(self):
        return Employee.objects.select_related(
            'user', 'bureau__agence__mutuelle', 'direction', 'manager__user'
        ).prefetch_related('contracts', 'documents', 'reports__user')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        # RH / DG : toujours autorisé
        if request.user.has_global_access:
            return super().dispatch(request, *args, **kwargs)
        # Directeur : autorisé si l'employé appartient à son périmètre
        if request.user.is_directeur:
            from apps.reporting.services import get_directeur_employees
            target = self.get_object()
            if get_directeur_employees(request.user).filter(pk=target.pk).exists():
                return super().dispatch(request, *args, **kwargs)
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Vous n'avez pas accès à la fiche de cet employé.")


# ============ Création ============

class EmployeeCreateView(GlobalAccessRequiredMixin, FormView):
    template_name = 'employees/employee_form.html'
    form_class = EmployeeCreateForm
    success_url = reverse_lazy('employees:employee_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['mode'] = 'create'
        ctx['position_suggestions'] = (
            Employee.objects.exclude(position='').values_list('position', flat=True).distinct().order_by('position')
        )
        return ctx

    def form_valid(self, form):
        with transaction.atomic():
            employee = form.save()
        messages.success(
            self.request,
            f"Employé créé : {employee.user.matricule} — {employee.user.get_full_name()}."
        )
        return redirect('employees:employee_detail', pk=employee.pk)


# ============ Édition ============

class EmployeeUpdateView(GlobalAccessRequiredMixin, FormView):
    template_name = 'employees/employee_form.html'
    form_class = EmployeeUpdateForm

    def get_employee(self):
        return get_object_or_404(Employee, pk=self.kwargs['pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['employee'] = self.get_employee()
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['mode'] = 'update'
        ctx['employee'] = self.get_employee()
        ctx['object'] = ctx['employee']
        ctx['position_suggestions'] = (
            Employee.objects.exclude(position='').values_list('position', flat=True).distinct().order_by('position')
        )
        return ctx

    def form_valid(self, form):
        with transaction.atomic():
            employee = form.save()
        messages.success(self.request, f"Employé {employee.user.matricule} mis à jour.")
        return redirect('employees:employee_detail', pk=employee.pk)


# ============ Suppression ============

class EmployeeDeleteView(GlobalAccessRequiredMixin, DeleteView):
    model = Employee
    template_name = 'employees/employee_confirm_delete.html'
    success_url = reverse_lazy('employees:employee_list')

    def form_valid(self, form):
        emp = self.get_object()
        # Soft delete plutôt que suppression physique
        emp.is_active = False
        emp.is_deleted = True
        emp.user.is_active = False
        emp.user.save()
        emp.save()
        messages.warning(self.request, f"Employé {emp.user.matricule} désactivé.")
        return redirect(self.success_url)


# ============ Mon profil (vue agent personnelle) ============

class MyProfileView(LoginRequiredMixin, View):
    """Vue de profil pour l'utilisateur connecté lui-même.

    Affiche : identité, affectation, planning, statistiques annuelles,
    et section de sécurité (changement de mot de passe).
    """
    template_name = 'employees/my_profile.html'

    def get(self, request):
        try:
            employee = request.user.employee
        except Employee.DoesNotExist:
            messages.warning(request, "Votre compte n'est pas rattaché à une fiche employé.")
            return redirect('core:home')

        from datetime import date, timedelta
        from apps.attendance.models import TimeEntry
        from apps.planning.services import get_active_planning
        import json

        planning = get_active_planning()
        daily_schedules = list(planning.schedules.order_by('day_of_week'))

        # Statistiques annuelles (heures par mois)
        # minutes = source de vérité (entier, cohérent avec les KPIs).
        today = date.today()
        year_chart = []
        for month in range(1, today.month + 1):
            entries = TimeEntry.objects.filter(
                employee=employee, work_date__year=today.year, work_date__month=month,
            )
            hours = timedelta()
            for e in entries:
                d = e.worked_duration
                if d:
                    hours += d
            year_chart.append({
                'label': ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin',
                          'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc'][month - 1],
                'minutes': int(hours.total_seconds() // 60),
            })

        return render(request, self.template_name, {
            'employee': employee,
            'planning': planning,
            'daily_schedules': daily_schedules,
            'year_chart_data': json.dumps(year_chart),
        })


# ============ Import en masse depuis Excel ============

DEFAULT_IMPORT_PASSWORD = 'Acep@2026'


class EmployeeImportView(GlobalAccessRequiredMixin, View):
    """Permet à la RH d'importer en masse des employés depuis un fichier Excel.

    Colonnes attendues : matricule | prenom | nom | direction | bureau | date_naissance | date_embauche
    Identification : direction et bureau peuvent être donnés par code OU par nom.
    Mot de passe initial pour tous les nouveaux comptes : 'Acep@2026'.
    Les employés déjà existants (matricule trouvé) sont ignorés.
    """
    template_name = 'employees/employee_import.html'

    def get(self, request):
        return render(request, self.template_name, {
            'default_password': DEFAULT_IMPORT_PASSWORD,
            'columns': ['matricule', 'prenom', 'nom', 'direction', 'bureau', 'date_naissance', 'date_embauche'],
        })

    def post(self, request):
        import openpyxl
        f = request.FILES.get('file')
        if not f:
            messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx).")
            return redirect('employees:employee_import')

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            messages.error(request, f"Impossible de lire le fichier : {e}")
            return redirect('employees:employee_import')

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "Le fichier est vide.")
            return redirect('employees:employee_import')

        # En-tête (1re ligne) — on tolère les variantes
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        idx = {}
        aliases = {
            'matricule': ['matricule', 'mat', 'id'],
            'prenom': ['prenom', 'prénom', 'first_name', 'firstname'],
            'nom': ['nom', 'last_name', 'lastname', 'name'],
            'direction': ['direction', 'service', 'dir'],
            'bureau': ['bureau', 'agence', 'site', 'bureau_affecté', 'bureau_affecte'],
            'date_naissance': ['date_naissance', 'naissance', 'birth_date', 'date naissance', 'datenaissance'],
            'date_embauche': ['date_embauche', 'embauche', 'hire_date', 'date embauche', 'dateembauche'],
        }
        for key, opts in aliases.items():
            for o in opts:
                if o in headers:
                    idx[key] = headers.index(o)
                    break

        missing = [k for k in ('matricule', 'prenom', 'nom') if k not in idx]
        if missing:
            messages.error(request, f"Colonnes obligatoires manquantes : {', '.join(missing)}.")
            return redirect('employees:employee_import')

        created, skipped, errors = 0, 0, []

        for row_num, row in enumerate(rows[1:], start=2):
            matricule = str(row[idx['matricule']]).strip() if row[idx['matricule']] else ''
            if not matricule:
                continue

            if User.objects.filter(matricule=matricule).exists():
                skipped += 1
                continue

            prenom = str(row[idx['prenom']]).strip() if row[idx['prenom']] else ''
            nom = str(row[idx['nom']]).strip() if row[idx['nom']] else ''

            # Direction (code ou nom)
            direction = None
            if 'direction' in idx and row[idx['direction']]:
                d_val = str(row[idx['direction']]).strip()
                direction = (
                    Direction.objects.filter(code__iexact=d_val).first()
                    or Direction.objects.filter(name__iexact=d_val).first()
                )

            # Bureau (code ou nom)
            bureau = None
            if 'bureau' in idx and row[idx['bureau']]:
                b_val = str(row[idx['bureau']]).strip()
                bureau = (
                    Bureau.objects.filter(code__iexact=b_val).first()
                    or Bureau.objects.filter(name__iexact=b_val).first()
                )

            if not direction or not bureau:
                errors.append(f"Ligne {row_num} ({matricule}) : direction ou bureau introuvable.")
                continue

            # Dates
            def _parse_date(v):
                if v is None or v == '':
                    return None
                if isinstance(v, datetime):
                    return v.date()
                if isinstance(v, date):
                    return v
                s = str(v).strip()
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except ValueError:
                        continue
                return None

            birth_date = _parse_date(row[idx['date_naissance']]) if 'date_naissance' in idx else None
            hire_date = _parse_date(row[idx['date_embauche']]) if 'date_embauche' in idx else date.today()
            if not hire_date:
                hire_date = date.today()

            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        matricule=matricule,
                        email=f"{matricule}@acep.sn",
                        password=DEFAULT_IMPORT_PASSWORD,
                        first_name=prenom,
                        last_name=nom,
                    )
                    user.roles = ['AGENT']
                    user.must_change_password = True
                    user.save()

                    Employee.objects.create(
                        user=user,
                        birth_date=birth_date,
                        hire_date=hire_date,
                        bureau=bureau,
                        direction=direction,
                        position='À renseigner',
                    )
                created += 1
            except Exception as e:
                errors.append(f"Ligne {row_num} ({matricule}) : {e}")

        if created:
            messages.success(request, f"{created} employé(s) créé(s). Mot de passe initial : {DEFAULT_IMPORT_PASSWORD} (à changer à la 1ère connexion).")
        if skipped:
            messages.info(request, f"{skipped} matricule(s) déjà existant(s) — ignoré(s).")
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f"... et {len(errors) - 10} autre(s) erreur(s).")

        return redirect('employees:employee_list')


class EmployeeImportTemplateView(GlobalAccessRequiredMixin, View):
    """Télécharge un modèle Excel vide avec les bonnes colonnes."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employés ACEP'

        headers = ['matricule', 'prenom', 'nom', 'direction', 'bureau', 'date_naissance', 'date_embauche']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='02564A', end_color='02564A', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Ligne d'exemple
        example = ['1042', 'Aliou', 'Niang', 'DIR-COM', 'BUR-VDN-01', '1990-05-15', '2020-01-15']
        for col, v in enumerate(example, 1):
            ws.cell(row=2, column=col, value=v)

        # Notes en bas
        ws.cell(row=4, column=1, value="Notes :").font = Font(bold=True)
        ws.cell(row=5, column=1, value="• direction et bureau peuvent être donnés par CODE ou par NOM.")
        ws.cell(row=6, column=1, value="• date_naissance et date_embauche au format YYYY-MM-DD ou DD/MM/YYYY.")
        ws.cell(row=7, column=1, value="• Le mot de passe initial sera Acep@2026 (à changer au 1er login).")
        ws.cell(row=8, column=1, value="• Les autres infos (rôles, poste, etc.) se modifient ensuite via la fiche employé.")

        for col, w in enumerate([12, 15, 15, 18, 18, 16, 16], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="modele_import_employes.xlsx"'
        wb.save(response)
        return response


